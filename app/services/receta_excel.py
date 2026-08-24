"""
Actualiza el flag de receta del catálogo desde el Excel "Base Predictiva de
Stock" — trae un cruce de receta por SKU individual (columna "Requiere Receta
(cruce SKU)"), mucho más confiable que la heurística por categoría/whitelist
que usa el catálogo por defecto.

Decisión (21/8): la whitelist manual de venta libre (sku_service.VENTA_LIBRE,
confirmada por la farmacia en la minuta del 31/7) queda BLINDADA. El Excel
suma información nueva donde no había, pero no la pisa: al correrlo contra el
catálogo real, 33 de sus "con receta" caían justo sobre esa whitelist —
Ibupirac, Actron, Tafirol, Aspirineta, Mylanta — los OTC más comunes de
Argentina. La hoja "Resumen" del Excel documenta en detalle cómo se corrigió
el precio, pero no dice nada de cómo se armó el cruce de receta: sin ese
método a la vista, no hay base para pisar un dato ya confirmado a mano. Cada
conflicto se reporta igual, para que Belén los revise caso por caso.

Los productos "No encontrado en cruce" (sin dato confiable) NO se tocan: el
catálogo sigue calculando su flag como hasta ahora.

Solo actualiza `requiere_receta`. No toca nombre, precio, stock ni categoría
— para eso está `catalogo_pdf.fusionar_con_catalogo`.
"""

import csv
import io
import logging

logger = logging.getLogger(__name__)

COLUMNAS_B = [
    "sku_id", "barcode", "sku_nombre", "sku_nombre_original", "marca",
    "laboratorio", "categoria", "es_medicamento", "precio_venta",
    "stock_actual", "ventas_mes", "prom_semanal", "cantidad_visible",
    "tipo_producto", "pausado", "requiere_receta", "clasificacion",
]

# Las 4 hojas del Excel (una por rubro) comparten el mismo layout de columnas.
_HOJAS = ("Alimentos", "Cosmeticos", "General", "Medicamentos")
_COL_BARCODE = 2   # "Codigo de barra" (0-based)
_COL_RECETA = 15   # "Requiere Receta (cruce SKU)" (0-based)

# Placeholder de código de barras faltante en el informe fuente: "1" no
# identifica un producto real, y varios productos distintos lo comparten —
# cruzarlo pisaría el flag de cualquiera de ellos al azar.
_BARCODE_PLACEHOLDER = "1"


def _normalizar(valor) -> str | None:
    """
    'si'/'Si' → 'si' · 'no'/'No aplica (no es medicamento)' → 'no' ·
    cualquier otra cosa (sin cruce, vacío, ruido) → None (no se toca).
    """
    if not isinstance(valor, str):
        return None
    v = valor.strip().lower()
    if v == "si":
        return "si"
    if v == "no" or v.startswith("no aplica"):
        return "no"
    return None


def parsear_excel(contenido: bytes) -> dict[str, str]:
    """
    Devuelve {barcode: 'si'|'no'} para todos los productos con cruce
    confiable, recorriendo las 4 hojas de rubro. Barcodes placeholder o sin
    cruce quedan afuera del diccionario (no se tocan al fusionar).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    resultado: dict[str, str] = {}
    for nombre_hoja in _HOJAS:
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        filas = ws.iter_rows(min_row=2, values_only=True)
        for row in filas:
            if len(row) <= max(_COL_BARCODE, _COL_RECETA):
                continue
            barcode = row[_COL_BARCODE]
            barcode = str(barcode).strip() if barcode is not None else ""
            if not barcode or barcode == _BARCODE_PLACEHOLDER:
                continue
            receta = _normalizar(row[_COL_RECETA])
            if receta is not None:
                resultado[barcode] = receta
    return resultado


def fusionar_receta(catalogo_actual: list, por_barcode_excel: dict[str, str]) -> tuple[str, dict]:
    """
    Actualiza SOLO `requiere_receta`, cruzando por código de barras. Todo lo
    demás del catálogo queda intacto — incluidos los productos sin cruce en
    el Excel, que conservan lo que ya tenían.

    `catalogo_actual`: lista de SKU (SKUService.todos()).
    Devuelve (csv_completo, resumen) con la cuenta de actualizados y la lista
    de conflictos con la whitelist de venta libre, para revisión manual.
    """
    from app.services.sku_service import es_venta_libre

    filas: dict[str, dict] = {}
    actualizados = 0
    conflictos: list[dict] = []

    for sku in catalogo_actual:
        nueva_receta = por_barcode_excel.get(sku.barcode)
        receta_final = sku.requiere_receta
        if nueva_receta is not None and nueva_receta != sku.requiere_receta:
            if nueva_receta == "si" and es_venta_libre(sku.sku_nombre):
                # Whitelist confirmada por la farmacia: no se pisa. Se
                # reporta para que Belén decida caso por caso.
                conflictos.append({
                    "barcode": sku.barcode, "nombre": sku.sku_nombre,
                    "antes": sku.requiere_receta, "excel": nueva_receta,
                })
            else:
                receta_final = nueva_receta
                actualizados += 1

        filas[sku.barcode] = {
            "sku_id": sku.sku_id, "barcode": sku.barcode,
            "sku_nombre": sku.sku_nombre, "sku_nombre_original": sku.sku_nombre_original,
            "marca": sku.marca, "laboratorio": sku.laboratorio, "categoria": sku.categoria,
            "es_medicamento": "si" if sku.es_medicamento else "no",
            "precio_venta": f"{sku.precio_venta:.2f}",
            "stock_actual": "" if sku.stock_actual is None else f"{sku.stock_actual:g}",
            "ventas_mes": "" if sku.ventas_mes is None else f"{sku.ventas_mes:.1f}",
            "prom_semanal": "" if sku.prom_semanal is None else f"{sku.prom_semanal:.2f}",
            "cantidad_visible": f"{sku.cantidad_visible}",
            "tipo_producto": sku.tipo_producto, "pausado": str(sku.pausado),
            "requiere_receta": receta_final, "clasificacion": sku.clasificacion,
        }

    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNAS_B, lineterminator="\n")
    w.writeheader()
    for fila in filas.values():
        w.writerow(fila)

    resumen = {
        "Total en el catálogo": len(catalogo_actual),
        "Con cruce en el Excel": len(por_barcode_excel),
        "Actualizados": actualizados,
        "Conflictos con la whitelist (no aplicados)": len(conflictos),
        "conflictos": conflictos,
    }
    return buf.getvalue(), resumen
